#!/usr/bin/env python3
"""
Bulk-import the historical paid-campaign leads in `All Leads.xlsx` into
wacrm. Calls the existing `ingest_lead` RPC (atomic, dedupe-by-last-10),
so re-running is safe — every row that's already in `contacts` matches
by phone and is skipped.

Three sheets, three schemas:
  Sheet1: Name, Phone, email   (4687 rows)
  Sheet4: Name, Email, Number  (2235 rows)
  Sheet6: Name, Number         (499 rows)

Every imported contact gets:
  - source:fb-lead         (closest existing source tag for paid campaigns)
  - import:past-campaigns  (added after the RPC — lets you filter / sweep
                            this batch later)
  - stage:new              (the RPC adds this unconditionally)

We deliberately do NOT add program:cohort / program:d2d — these leads
weren't qualified for either programme, and the nurture automations
listen on those tags. Importing with program:* would mass-fire the drip
the moment templates flip to APPROVED. That decision is yours, not the
import's.

Env needed (read from wacrm/.env.local automatically if present):
  NEXT_PUBLIC_SUPABASE_URL
  SUPABASE_SERVICE_ROLE_KEY
  LEAD_INGEST_USER_ID

Usage:
  /tmp/.venv-xlsx/bin/python scripts/import-leads-xlsx.py [--dry-run] [--xlsx PATH]
"""
import argparse
import json
import os
import re
import ssl
import sys
import time
import urllib.request
import urllib.error
from pathlib import Path

import certifi
import openpyxl

# python.org Python on macOS ships without a system cert bundle — point
# urllib at certifi's CA bundle so HTTPS to Supabase verifies.
SSL_CTX = ssl.create_default_context(cafile=certifi.where())


# ----------------------------------------------------------------------
# Env loading (no python-dotenv dependency — parse .env.local ourselves)
# ----------------------------------------------------------------------
def load_dotenv(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith('#') or '=' not in line:
            continue
        k, v = line.split('=', 1)
        v = v.strip().strip('"').strip("'")
        os.environ.setdefault(k.strip(), v)


# ----------------------------------------------------------------------
# Phone canonicalisation — mirrors src/lib/leads/ingest.ts so the import
# format matches what the website + webhook produce. Returns the 12-digit
# `91XXXXXXXXXX` form on success, or None to skip.
# ----------------------------------------------------------------------
def canonicalize_phone(raw) -> str | None:
    if raw is None:
        return None
    # Excel stores numeric phones as float — drop the trailing .0
    if isinstance(raw, float):
        raw = str(int(raw))
    s = re.sub(r'\D', '', str(raw))
    if not s:
        return None
    # 10-digit mobile → prefix 91
    if len(s) == 10 and s[0] in '6789':
        s = '91' + s
    # 0 + 10-digit (trunk-prefixed) → drop 0, prefix 91
    elif len(s) == 11 and s[0] == '0' and s[1] in '6789':
        s = '91' + s[1:]
    # 91 + 0 + 10-digit → drop the trunk 0
    elif len(s) == 13 and s.startswith('910'):
        s = '91' + s[3:]
    # else: keep whatever it is (incl. already-12-digit 91…)
    # Reject anything that isn't a plausible E.164 length.
    if len(s) < 10 or len(s) > 15:
        return None
    return s


def clean_email(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or '@' not in s:
        return None
    return s


def clean_name(raw) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip()
    return s or None


# ----------------------------------------------------------------------
# Per-sheet readers — return list of (name, phone, email).
# ----------------------------------------------------------------------
SHEET_SCHEMAS = {
    # sheet_name: (name_col_idx, phone_col_idx, email_col_idx_or_None)
    'Sheet1': (0, 1, 2),
    'Sheet4': (0, 2, 1),
    'Sheet6': (0, 1, None),
}


def read_rows(xlsx_path: Path):
    """
    Yields (sheet_name, row_idx, name, phone_raw, email) for every data row.

    Self-heals two real defects in the source spreadsheet:
      * The Phone / Email columns are swapped starting at Sheet1 row 3320
        (verified: emails landed in the phone column, phones landed in
        the email column). When the declared phone cell doesn't look like
        a phone but another cell on the row does, we use that cell as
        the phone — and treat the original phone cell as the email.
      * Excel auto-converts >15-digit numbers to scientific notation
        (`9.19...e+16`); those are unrecoverable and stay rejected.
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    for sheet_name in wb.sheetnames:
        schema = SHEET_SCHEMAS.get(sheet_name)
        if not schema:
            print(f'  [skip] unknown sheet schema: {sheet_name}')
            continue
        name_i, phone_i, email_i = schema
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                continue  # header
            if not row or all(c is None or str(c).strip() == '' for c in row):
                continue
            name = clean_name(row[name_i]) if name_i < len(row) else None
            phone_raw = row[phone_i] if phone_i < len(row) else None
            email = clean_email(row[email_i]) if email_i is not None and email_i < len(row) else None

            # Column-swap recovery: if the declared phone cell doesn't
            # canonicalize but a different cell on the row does, swap them.
            if not canonicalize_phone(phone_raw):
                for ci, candidate in enumerate(row):
                    if ci in (name_i, phone_i):
                        continue
                    if canonicalize_phone(candidate):
                        # The original phone cell is probably the email.
                        recovered_email = clean_email(phone_raw)
                        if recovered_email and not email:
                            email = recovered_email
                        phone_raw = candidate
                        break

            yield sheet_name, row_idx, name, phone_raw, email


# ----------------------------------------------------------------------
# Supabase REST/RPC client (urllib — no extra deps).
# ----------------------------------------------------------------------
class SupabaseRPC:
    def __init__(self, url: str, service_key: str):
        self.url = url.rstrip('/')
        self.service_key = service_key

    def _headers(self) -> dict:
        return {
            'apikey': self.service_key,
            'Authorization': f'Bearer {self.service_key}',
            'Content-Type': 'application/json',
            'Prefer': 'return=representation',
        }

    def rpc(self, fn: str, args: dict):
        req = urllib.request.Request(
            f'{self.url}/rest/v1/rpc/{fn}',
            data=json.dumps(args).encode(),
            headers=self._headers(),
            method='POST',
        )
        try:
            with urllib.request.urlopen(req, timeout=30, context=SSL_CTX) as resp:
                body = resp.read().decode()
                return json.loads(body) if body else None
        except urllib.error.HTTPError as e:
            err_body = e.read().decode()
            raise RuntimeError(f'RPC {fn} failed: {e.code} {err_body}') from e

    def get(self, path: str, params: dict | None = None):
        q = ''
        if params:
            q = '?' + '&'.join(f'{k}={v}' for k, v in params.items())
        req = urllib.request.Request(
            f'{self.url}/rest/v1/{path}{q}',
            headers=self._headers(),
            method='GET',
        )
        with urllib.request.urlopen(req, timeout=30, context=SSL_CTX) as resp:
            return json.loads(resp.read().decode())

    def post(self, path: str, body: dict | list):
        req = urllib.request.Request(
            f'{self.url}/rest/v1/{path}',
            data=json.dumps(body).encode(),
            headers={**self._headers(), 'Prefer': 'resolution=ignore-duplicates'},
            method='POST',
        )
        try:
            with urllib.request.urlopen(req, timeout=30, context=SSL_CTX) as resp:
                return json.loads(resp.read().decode() or '[]')
        except urllib.error.HTTPError as e:
            # Conflict on the unique key is the idempotent path — treat as no-op.
            if e.code in (409,):
                return []
            raise RuntimeError(f'POST {path} failed: {e.code} {e.read().decode()}') from e


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--xlsx', default='All Leads.xlsx', help='Path to All Leads.xlsx (relative to wacrm/).')
    ap.add_argument('--dry-run', action='store_true', help='Parse + canonicalise but don\'t hit Supabase.')
    ap.add_argument('--limit', type=int, default=0, help='Only process the first N rows (0 = all).')
    args = ap.parse_args()

    wacrm_root = Path(__file__).resolve().parent.parent
    load_dotenv(wacrm_root / '.env.local')

    xlsx_path = wacrm_root / args.xlsx
    if not xlsx_path.exists():
        print(f'ERROR: {xlsx_path} not found', file=sys.stderr)
        sys.exit(1)

    supabase_url = os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
    service_key = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
    user_id = os.environ.get('LEAD_INGEST_USER_ID')
    if not args.dry_run and (not supabase_url or not service_key or not user_id):
        print('ERROR: NEXT_PUBLIC_SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY / LEAD_INGEST_USER_ID missing', file=sys.stderr)
        sys.exit(1)

    client = None if args.dry_run else SupabaseRPC(supabase_url, service_key)

    # Resolve / create the "import:past-campaigns" tag once, before the loop.
    import_tag_id = None
    if not args.dry_run:
        existing = client.get(
            'tags',
            {'user_id': f'eq.{user_id}', 'name': 'eq.import:past-campaigns', 'select': 'id'},
        )
        if existing:
            import_tag_id = existing[0]['id']
        else:
            created = client.post('tags', {
                'user_id': user_id,
                'name': 'import:past-campaigns',
                'color': '#94a3b8',
            })
            import_tag_id = created[0]['id'] if created else client.get(
                'tags', {'user_id': f'eq.{user_id}', 'name': 'eq.import:past-campaigns', 'select': 'id'},
            )[0]['id']
        print(f'  import tag id: {import_tag_id}')

    stats = {
        'total': 0,
        'skipped_no_phone': 0,
        'skipped_bad_phone': 0,
        'ingested': 0,
        'deduped': 0,
        'errors': 0,
        'by_sheet': {},
    }

    start = time.time()
    for sheet_name, row_idx, name, phone_raw, email in read_rows(xlsx_path):
        stats['total'] += 1
        stats['by_sheet'].setdefault(sheet_name, 0)
        stats['by_sheet'][sheet_name] += 1

        if args.limit and stats['total'] > args.limit:
            break

        if phone_raw is None:
            stats['skipped_no_phone'] += 1
            continue
        phone = canonicalize_phone(phone_raw)
        if not phone:
            stats['skipped_bad_phone'] += 1
            if stats['skipped_bad_phone'] <= 10:
                print(f'  [skip bad phone] {sheet_name} row {row_idx}: {phone_raw!r}')
            continue

        if args.dry_run:
            continue

        # external_ref is per-sheet+last10 — makes re-running the script
        # idempotent on the import side too (the RPC dedupes by phone
        # last-10 anyway, but this keeps the cf clean).
        ext_ref = f'xlsx:{sheet_name.lower()}:{phone[-10:]}'
        try:
            result = client.rpc('ingest_lead', {
                'p_user_id': user_id,
                'p_phone': phone,
                'p_name': name or '',
                'p_email': email or '',
                'p_source': 'fb-lead',
                'p_program': '',
                'p_business_stage': '',
                'p_application_status': '',
                'p_portfolio': '',
                'p_marketing_consent': True,  # privacy-policy consent basis
                'p_external_ref': ext_ref,
                'p_notes': '',
                'p_create_deal': False,
            })
        except Exception as e:
            stats['errors'] += 1
            if stats['errors'] <= 10:
                print(f'  [error] {sheet_name} row {row_idx} phone {phone}: {e}')
            continue

        contact_id = result.get('contact_id') if isinstance(result, dict) else None
        if isinstance(result, dict) and result.get('deduped'):
            stats['deduped'] += 1
        else:
            stats['ingested'] += 1

        # Tag with import:past-campaigns. ON CONFLICT DO NOTHING via the
        # client's ignore-duplicates preference; harmless if already linked.
        if contact_id and import_tag_id:
            try:
                client.post('contact_tags', {
                    'contact_id': contact_id,
                    'tag_id': import_tag_id,
                })
            except Exception as e:
                # Don't count tag-link failures as ingest errors — the row
                # is still in contacts. Just log the first few.
                if stats['errors'] <= 10:
                    print(f'  [tag link warn] {sheet_name} row {row_idx}: {e}')

        if stats['total'] % 250 == 0:
            elapsed = time.time() - start
            rate = stats['total'] / max(elapsed, 0.001)
            print(f'  ... {stats["total"]} processed ({rate:.1f}/s, {elapsed:.0f}s elapsed)')

    elapsed = time.time() - start
    print('\n=== Done ===')
    print(json.dumps(stats, indent=2))
    print(f'elapsed: {elapsed:.1f}s')


if __name__ == '__main__':
    main()
